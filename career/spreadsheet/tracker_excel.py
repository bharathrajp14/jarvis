# career/spreadsheet/tracker_excel.py — 10-Sheet Master Career Tracker Excel Workbook
from __future__ import annotations

import logging
import time
from pathlib import Path
from typing import Any, Dict, List, Optional

try:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    _OPENPYXL_AVAILABLE = True
except ImportError:
    _OPENPYXL_AVAILABLE = False

from career.models import (
    Application,
    CareerContact,
    EmailEventRecord,
    FollowupRecord,
    InterviewSchedule,
    OfferCandidate,
)

logger = logging.getLogger("JARVIS.CareerSpreadsheet.TrackerExcel")


class CareerTrackerWorkbook:
    """
    Constructs and styles the authoritative 10-Sheet BR JARVIS Career Tracker Excel workbook.
    """

    # Headers for each of the 10 sheets
    SHEET_SCHEMAS = {
        "Applications": [
            "Application ID", "Company", "Role", "Location", "Job URL", "Source",
            "Platform", "Match %", "Resume Version", "Cover Letter", "Application Method",
            "Applied Date", "Application Status", "Submission Status", "Confirmation ID",
            "Next Follow-up", "Last Updated", "Priority", "Notes"
        ],
        "Jobs": [
            "Job ID", "Company", "Role", "Location", "Remote Type", "Salary",
            "Posted Date", "Source", "URL", "Match Score", "Status",
            "Discovered Date", "Last Checked"
        ],
        "Interviews": [
            "Interview ID", "Application ID", "Company", "Role", "Round", "Date",
            "Time", "Timezone", "Meeting URL", "Interviewer", "Status",
            "Preparation Status", "Notes"
        ],
        "Offers": [
            "Offer ID", "Application ID", "Company", "Role", "Salary", "Currency",
            "Bonus", "Benefits", "Location", "Work Mode", "Joining Date",
            "Offer Date", "Expiry Date", "Status", "Evidence", "Notes"
        ],
        "Followups": [
            "Followup ID", "Application ID", "Company", "Role", "Reason",
            "Due Date", "Priority", "Status", "Completed Date", "Notes"
        ],
        "Contacts": [
            "Contact ID", "Application ID", "Company", "Name", "Title",
            "Email", "Phone", "LinkedIn", "Notes"
        ],
        "Email Events": [
            "Email Event ID", "Application ID", "Message ID Hash", "Provider",
            "Sender", "Sender Domain", "Subject", "Received Time",
            "Classification", "Confidence", "Detected Event", "Action Taken", "Processed Time"
        ],
        "Resume Versions": [
            "Resume Version ID", "Template", "Role", "Job ID", "Created Date",
            "ATS Score", "File Path", "Status"
        ],
        "Analytics": [
            "Metric / KPI", "Value", "Benchmark Target", "Status / Notes"
        ],
        "Settings": [
            "Setting Key", "Setting Value", "Description", "Last Updated"
        ],
    }

    @classmethod
    def build_workbook(
        cls,
        applications: List[Application],
        interviews: List[InterviewSchedule],
        offers: List[OfferCandidate],
        followups: List[FollowupRecord],
        email_events: List[EmailEventRecord],
        contacts: List[CareerContact],
        analytics_summary: Optional[Dict[str, Any]] = None,
    ) -> Any:
        """Create styled 10-sheet openpyxl Workbook populated with database records."""
        if not _OPENPYXL_AVAILABLE:
            raise RuntimeError("openpyxl is required for Excel spreadsheet generation.")

        wb = openpyxl.Workbook()
        # Remove default active sheet to create tabs in deterministic order
        default_ws = wb.active

        # Standard Theme Styles
        header_fill = PatternFill(start_color="1A365D", end_color="1A365D", fill_type="solid")  # Deep Navy
        header_font = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
        data_font = Font(name="Segoe UI", size=9)
        title_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
        title_font = Font(name="Segoe UI", size=11, bold=True, color="00F2FE")

        thin_side = Side(style="thin", color="E2E8F0")
        thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

        for sheet_idx, (sheet_name, headers) in enumerate(cls.SHEET_SCHEMAS.items()):
            if sheet_idx == 0:
                ws = default_ws
                ws.title = sheet_name
            else:
                ws = wb.create_sheet(title=sheet_name)

            # 1. Write Header Row
            ws.append(headers)
            for col_num in range(1, len(headers) + 1):
                cell = ws.cell(row=1, column=col_num)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # 2. Populate Sheet Rows
            rows_data: List[List[Any]] = []
            if sheet_name == "Applications":
                for app in applications:
                    rows_data.append([
                        app.application_id,
                        app.company,
                        app.job_title,
                        app.location,
                        app.job_url,
                        app.source,
                        app.platform,
                        f"{app.match_score:.0f}%" if app.match_score else "N/A",
                        app.resume_version,
                        "Yes" if app.cover_letter_version else "No",
                        app.application_method,
                        app.date_applied or "Pending",
                        app.application_status.value if hasattr(app.application_status, "value") else str(app.application_status),
                        app.submission_status,
                        app.confirmation_id or "N/A",
                        app.next_followup or "None",
                        time.strftime("%Y-%m-%d %H:%M", time.localtime(app.last_updated)),
                        app.priority.value if hasattr(app.priority, "value") else str(app.priority),
                        "; ".join(app.notes[-2:]) if app.notes else "",
                    ])

            elif sheet_name == "Interviews":
                for iv in interviews:
                    rows_data.append([
                        iv.interview_id,
                        iv.application_id,
                        iv.company,
                        iv.role,
                        iv.round,
                        iv.date,
                        iv.time_str,
                        iv.timezone,
                        iv.meeting_url,
                        iv.interviewer,
                        iv.status,
                        iv.preparation_status,
                        "; ".join(iv.notes) if iv.notes else "",
                    ])

            elif sheet_name == "Offers":
                for off in offers:
                    rows_data.append([
                        off.offer_id,
                        off.application_id,
                        off.company,
                        off.role,
                        off.salary,
                        off.currency,
                        off.bonus or "N/A",
                        ", ".join(off.benefits) if off.benefits else "Standard Package",
                        off.location,
                        off.work_mode,
                        off.joining_date or "TBD",
                        off.offer_date,
                        off.expiry_date or "N/A",
                        off.status.value if hasattr(off.status, "value") else str(off.status),
                        off.evidence,
                        "; ".join(off.notes) if off.notes else "",
                    ])

            elif sheet_name == "Followups":
                for fol in followups:
                    rows_data.append([
                        fol.followup_id,
                        fol.application_id,
                        fol.company,
                        fol.role,
                        fol.reason,
                        fol.due_date,
                        fol.priority.value if hasattr(fol.priority, "value") else str(fol.priority),
                        fol.status,
                        fol.completed_date or "Pending",
                        "; ".join(fol.notes) if fol.notes else "",
                    ])

            elif sheet_name == "Contacts":
                for cnt in contacts:
                    rows_data.append([
                        cnt.contact_id,
                        cnt.application_id or "General",
                        cnt.company,
                        cnt.name,
                        cnt.title,
                        cnt.email,
                        cnt.phone,
                        cnt.linkedin_url,
                        cnt.notes,
                    ])

            elif sheet_name == "Email Events":
                for eml in email_events:
                    rows_data.append([
                        eml.email_event_id,
                        eml.application_id or "N/A",
                        eml.message_id_hash[:12] + "...",
                        eml.provider,
                        eml.sender,
                        eml.sender_domain,
                        eml.subject,
                        eml.received_time,
                        eml.classification.value if hasattr(eml.classification, "value") else str(eml.classification),
                        f"{eml.confidence*100:.0f}%",
                        eml.detected_event,
                        eml.action_taken,
                        time.strftime("%Y-%m-%d %H:%M", time.localtime(eml.processed_time)),
                    ])

            elif sheet_name == "Analytics":
                summary = analytics_summary or {}
                rows_data.extend([
                    ["Total Active Applications", summary.get("total_applications_submitted", len(applications)), "Active Goal", "Canonical DB"],
                    ["Verified Applications", summary.get("total_applications_submitted", len(applications)), "100%", "Evidence-based"],
                    ["Scheduled Interviews", summary.get("total_interviews", len(interviews)), "Active Loops", "Calendar synced"],
                    ["Offer Candidates", summary.get("total_offers", len(offers)), "Active", "Verified"],
                    ["Response Rate", f"{summary.get('response_rate', 0.0)}%", "> 25%", "Funnel Rate"],
                    ["Interview Conversion Rate", f"{summary.get('interview_rate', 0.0)}%", "> 15%", "Funnel Rate"],
                    ["Offer Conversion Rate", f"{summary.get('offer_rate', 0.0)}%", "> 30%", "Funnel Rate"],
                    ["Last Sync Timestamp", time.strftime("%Y-%m-%d %H:%M:%S"), "Real-time", "Projection Synced"],
                ])

            elif sheet_name == "Settings":
                rows_data.extend([
                    ["candidate_name", "Bharath", "Primary Applicant Identity", time.strftime("%Y-%m-%d")],
                    ["default_currency", "USD", "Preferred Offer & Salary Currency", time.strftime("%Y-%m-%d")],
                    ["followup_delay_days", "6", "Automatic first follow-up schedule window", time.strftime("%Y-%m-%d")],
                    ["offer_auto_confirm", "FALSE (STRICT_HUMAN_APPROVAL)", "Safety interlock for offer state transitions", time.strftime("%Y-%m-%d")],
                    ["calendar_auto_sync", "TRUE", "Auto-schedule verified interviews with buffer checks", time.strftime("%Y-%m-%d")],
                    ["prompt_injection_defense", "ACTIVE (FAIL_CLOSED)", "Wraps untrusted emails & job descriptions", time.strftime("%Y-%m-%d")],
                ])

            # Append formatted rows
            for r_idx, row in enumerate(rows_data, start=2):
                ws.append(row)
                for col_idx in range(1, len(row) + 1):
                    cell = ws.cell(row=r_idx, column=col_idx)
                    cell.font = data_font
                    cell.border = thin_border

            # 3. Auto Column Sizing
            for col in ws.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    val_str = str(cell.value or "")
                    if len(val_str) > max_len:
                        max_len = len(val_str)
                ws.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 40)

        return wb
